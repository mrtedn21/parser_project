from openpyxl import load_workbook

from rest_framework.response import Response
from rest_framework.status import HTTP_200_OK
from rest_framework.mixins import CreateModelMixin
from rest_framework.mixins import ListModelMixin
from rest_framework.mixins import RetrieveModelMixin
from rest_framework.permissions import IsAuthenticated
from rest_framework.viewsets import GenericViewSet

from parser.models import Bill
from parser.serializers import BillSerializer
from parser.serializers import BillFileSerializer


class BillViewSet(GenericViewSet,
                  ListModelMixin,
                  RetrieveModelMixin):
    queryset = Bill.objects.all()
    serializer_class = BillSerializer
    permission_classes = [IsAuthenticated]


class BillsFileViewSet(GenericViewSet,
                       CreateModelMixin):
    queryset = Bill.objects.all()
    serializer_class = BillFileSerializer
    permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        file = request.FILES['file']
        workbook = load_workbook(file)
        worksheet = workbook.worksheets[0]

        for row in data_iterator(worksheet):
            serializer = BillSerializer(data=row)
            if serializer.is_valid():
                serializer.save()

        return Response(data={'result': 'ok'}, status=HTTP_200_OK)


def data_iterator(sheet):
    index = 2
    while True:
        if not sheet.cell(row=index, column=1).value:
            break
        yield {
            'client_name': sheet.cell(row=index, column=1).value,
            'client_org': sheet.cell(row=index, column=2).value,
            'number': str(sheet.cell(row=index, column=3).value),
            'sum': str(sheet.cell(row=index, column=4).value),
            'date': sheet.cell(row=index, column=5).value.strftime(
                '%Y-%m-%d'),
            'service': sheet.cell(row=index, column=6).value,
        }
        index += 1
